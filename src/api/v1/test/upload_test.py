from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Path
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import load_workbook
import io

from src.base.db import get_db
from src.model import Question, Subject

router = APIRouter()


@router.post("/upload-questions/{subject_id}")
async def upload_questions(
        subject_id: int = Path(..., description="Subject ID raqami", gt=0),
        file: UploadFile = File(..., description="Excel fayl (.xlsx, .xls)"),
        db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Faqat Excel fayllar qabul qilinadi (.xlsx, .xls)")

    subject = await db.scalar(select(Subject).where(Subject.id == subject_id))
    if not subject:
        raise HTTPException(status_code=404, detail=f"Subject ID {subject_id} topilmadi")

    try:
        contents = await file.read()
        workbook = load_workbook(filename=io.BytesIO(contents), data_only=True)
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel faylni o'qishda xatolik: {e}")

    added_count = 0
    errors = []

    try:
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Bo'sh qatorlarni o'tkazib yuborish
                if not row or all(cell is None or str(cell).strip() == '' for cell in row[:5]):
                    continue

                # Kamida 5 ta ustun borligini tekshirish (Savol + 4 ta javob)
                if len(row) < 5:
                    errors.append(f"Qator {row_num}: 5 ta ustun bo'lishi kerak (Savol + 4 ta javob)")
                    continue

                question_text = str(row[0]).strip() if row[0] is not None else ""
                option_a = str(row[1]).strip() if row[1] is not None else ""  # To'g'ri javob
                option_b = str(row[2]).strip() if row[2] is not None else ""
                option_c = str(row[3]).strip() if row[3] is not None else ""
                option_d = str(row[4]).strip() if row[4] is not None else ""

                # Bo'sh maydonlarni tekshirish
                if not all([question_text, option_a, option_b, option_c, option_d]):
                    errors.append(f"Qator {row_num}: Savol va barcha 4 ta javob to'ldirilishi kerak")
                    continue

                # To'g'ri javobni belgilash - birinchi variant (option_a) to'g'ri javob
                correct_option_text = option_a  # Har doim birinchi variant to'g'ri

                # Savolni yaratish va bazaga qo'shish - faqat text maydonlar
                question = Question(
                    text=question_text,
                    question_image="",  # Rasm yo'q
                    option_a=option_a,
                    option_a_image="",  # Rasm yo'q
                    option_b=option_b,
                    option_b_image="",  # Rasm yo'q
                    option_c=option_c,
                    option_c_image="",  # Rasm yo'q
                    option_d=option_d,
                    option_d_image="",  # Rasm yo'q
                    correct_option_text=correct_option_text,
                    correct_option_image="",  # Rasm yo'q
                    subject_id=subject_id
                )

                db.add(question)
                added_count += 1

            except Exception as e:
                errors.append(f"Qator {row_num}: {str(e)}")

        await db.commit()

    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Bazaga saqlashda xatolik: {str(e)}")

    return {
        "message": f"{added_count} ta savol muvaffaqiyatli yuklandi",
        "success_count": added_count,
        "errors": errors,
        "subject_id": subject_id
    }
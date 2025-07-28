from fastapi import APIRouter, UploadFile, File, Depends
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
import io

from src.base.db import get_db
from src.model import Question

router = APIRouter()

@router.post("/upload-questions/")
async def upload_questions(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    content = await file.read()
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        text, correct_text, a, b, c, d = row

        # Matnga qarab to‘g‘ri variant harfini aniqlaymiz
        correct_option = None
        if correct_text == a:
            correct_option = "A"
        elif correct_text == b:
            correct_option = "B"
        elif correct_text == c:
            correct_option = "C"
        elif correct_text == d:
            correct_option = "D"
        else:
            raise ValueError(f"{i}-qatorda to‘g‘ri javob variantlar bilan mos emas: {correct_text}")

        question = Question(
            text=text,
            option_a=a,
            option_b=b,
            option_c=c,
            option_d=d,
            correct_option=correct_option
        )
        db.add(question)

    await db.commit()
    return {"message": "Savollar muvaffaqiyatli yuklandi"}
